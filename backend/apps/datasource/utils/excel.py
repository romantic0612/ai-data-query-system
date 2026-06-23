import re
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook


EDUCATION_STATS_SHEET_NAME = "education_stats_long"

FIELD_TYPE_MAP = {
    'int64': 'int',
    'int32': 'int',
    'float64': 'float',
    'float32': 'float',
    'datetime64': 'datetime',
    'datetime64[ns]': 'datetime',
    'object': 'string',
    'string': 'string',
    'bool': 'string',
}

USER_TYPE_TO_PANDAS = {
    'int': 'int64',
    'float': 'float64',
    'datetime': 'datetime64[ns]',
    'string': 'string',
}

REPORT_CODE_LABEL = "\u8868\u53f7"
STAT_TIME_LABEL = "\u7edf\u8ba1\u65f6\u70b9"
DIRECTORY_LABEL = "\u76ee\u5f55"

CHINESE_CODE_MARKERS = {
    "\u7532", "\u4e59", "\u4e19", "\u4e01", "\u620a",
    "\u5df1", "\u5e9a", "\u8f9b", "\u58ec", "\u7678",
}
DIMENSION_TERMS = (
    "\u6307\u6807", "\u540d\u79f0", "\u4ee3\u7801", "\u4e13\u4e1a",
    "\u5206\u7c7b", "\u7c7b\u522b", "\u5e74\u5236", "\u5b66\u6821",
    "\u9662\u7cfb", "\u5355\u4f4d", "\u8ba1\u91cf\u5355\u4f4d",
)


def infer_field_type(dtype) -> str:
    dtype_str = str(dtype)
    return FIELD_TYPE_MAP.get(dtype_str, 'string')


def parse_excel_preview(save_path: str, max_rows: int = 10):
    sheets_data = []
    if save_path.endswith(".csv"):
        df = pd.read_csv(save_path, engine='c')
        fields = _build_fields(df)
        preview_data = _preview_records(df, max_rows)
        sheets_data.append({
            "sheetName": "Sheet1",
            "fields": fields,
            "data": preview_data,
            "rows": len(df)
        })
    else:
        education_df = parse_education_statistics_workbook(save_path)
        if not education_df.empty:
            sheets_data.append({
                "sheetName": EDUCATION_STATS_SHEET_NAME,
                "fields": _build_fields(education_df),
                "data": _preview_records(education_df, max_rows),
                "rows": len(education_df)
            })
            return sheets_data

        sheet_names = pd.ExcelFile(save_path).sheet_names
        for sheet_name in sheet_names:
            df = pd.read_excel(save_path, sheet_name=sheet_name, engine='calamine')
            fields = _build_fields(df)
            preview_data = _preview_records(df, max_rows)
            sheets_data.append({
                "sheetName": sheet_name,
                "fields": fields,
                "data": preview_data,
                "rows": len(df)
            })
    return sheets_data


def parse_education_statistics_workbook(save_path: str) -> pd.DataFrame:
    if save_path.lower().endswith((".csv", ".xls")):
        return pd.DataFrame()

    try:
        workbook = load_workbook(save_path, data_only=True, read_only=False)
    except Exception:
        return pd.DataFrame()

    records: List[Dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            if _is_directory_sheet(worksheet.title):
                continue

            matrix = _worksheet_matrix(worksheet)
            if not _looks_like_education_stats(matrix):
                continue

            meta = _report_meta(worksheet.title, matrix)
            code_row_index = _find_code_row(matrix)
            if code_row_index is not None:
                records.extend(_parse_cross_table(matrix, meta, code_row_index))
            else:
                records.extend(_parse_key_value_table(matrix, meta))
    finally:
        workbook.close()

    if not records:
        return pd.DataFrame()

    columns = [
        "source_sheet",
        "report_code",
        "report_name",
        "stat_period",
        "stat_year",
        "row_name",
        "row_code",
        "dimension_path",
        "dimension_1",
        "dimension_2",
        "dimension_3",
        "dimension_4",
        "metric_code",
        "metric_name",
        "metric_path",
        "metric_value",
        "metric_number",
        "unit",
    ]
    return pd.DataFrame(records, columns=columns)


def _build_fields(df: pd.DataFrame) -> List[Dict[str, str]]:
    return [
        {
            "fieldName": str(col),
            "fieldType": infer_field_type(df[col].dtype),
        }
        for col in df.columns
    ]


def _preview_records(df: pd.DataFrame, max_rows: int) -> List[Dict[str, Any]]:
    preview_df = df.head(max_rows).astype(object)
    preview_df = preview_df.where(pd.notnull(preview_df), None)
    return preview_df.to_dict(orient='records')


def _is_directory_sheet(sheet_name: str) -> bool:
    return DIRECTORY_LABEL in str(sheet_name)


def _worksheet_matrix(worksheet) -> List[List[Any]]:
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 0
    matrix = [
        [_clean_value(worksheet.cell(row=row, column=col).value) for col in range(1, max_col + 1)]
        for row in range(1, max_row + 1)
    ]

    for merged_range in worksheet.merged_cells.ranges:
        top_value = _clean_value(
            worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        )
        if top_value is None:
            continue
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if row <= max_row and col <= max_col:
                    matrix[row - 1][col - 1] = top_value
    return matrix


def _clean_value(value: Any) -> Optional[Any]:
    if value is None:
        return None
    if isinstance(value, str):
        text = re.sub(r"\s+", " ", value.replace("\u3000", " ")).strip()
        return text or None
    return value


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_education_stats(matrix: List[List[Any]]) -> bool:
    top_cells = [_text(cell) for row in matrix[:6] for cell in row[:12] if _text(cell)]
    if not top_cells:
        return False

    has_report_code = any(REPORT_CODE_LABEL in cell for cell in top_cells)
    has_stat_time = any(STAT_TIME_LABEL in cell for cell in top_cells)
    has_code_row = _find_code_row(matrix) is not None
    has_education_sheet_title = any(cell.startswith("\u6559\u57fa") for cell in top_cells)
    return (has_report_code or has_education_sheet_title) and (has_stat_time or has_code_row)


def _report_meta(sheet_name: str, matrix: List[List[Any]]) -> Dict[str, Optional[str]]:
    first_row_text = [_text(cell) for cell in matrix[0] if _text(cell)] if matrix else []
    report_name = first_row_text[0] if first_row_text else sheet_name
    report_code = None
    stat_period = None

    for row in matrix[:5]:
        for cell in row:
            text = _text(cell)
            if not text:
                continue
            if REPORT_CODE_LABEL in text and report_code is None:
                report_code = _value_after_label(text, REPORT_CODE_LABEL)
            if STAT_TIME_LABEL in text and stat_period is None:
                stat_period = _value_after_label(text, STAT_TIME_LABEL)

    if not report_code:
        match = re.search(r"(\u6559\u57fa[\w\u4e00-\u9fa5]+)", sheet_name)
        report_code = match.group(1) if match else None

    stat_year = None
    if stat_period:
        year_match = re.search(r"(\d{4})", stat_period)
        stat_year = year_match.group(1) if year_match else None

    return {
        "source_sheet": sheet_name,
        "report_code": report_code,
        "report_name": report_name,
        "stat_period": stat_period,
        "stat_year": stat_year,
    }


def _value_after_label(text: str, label: str) -> str:
    parts = re.split(r"[:\uff1a]", text, maxsplit=1)
    if len(parts) == 2:
        return parts[1].strip()
    return text.replace(label, "").strip()


def _find_code_row(matrix: List[List[Any]]) -> Optional[int]:
    for row_index, row in enumerate(matrix[:12]):
        values = {_text(value) for value in row if _text(value)}
        if "\u7532" in values and "\u4e59" in values:
            return row_index
        numeric_markers = sum(1 for value in values if re.fullmatch(r"\d+", value))
        chinese_markers = sum(1 for value in values if value in CHINESE_CODE_MARKERS)
        if numeric_markers >= 2 and chinese_markers >= 1:
            return row_index
    return None


def _parse_cross_table(
    matrix: List[List[Any]],
    meta: Dict[str, Optional[str]],
    code_row_index: int,
) -> List[Dict[str, Any]]:
    if not matrix:
        return []

    max_col = max((len(row) for row in matrix), default=0)
    header_start_index = 2 if code_row_index >= 3 else max(0, code_row_index - 1)
    columns = []

    for col_index in range(max_col):
        code = _cell(matrix, code_row_index, col_index)
        parts = _header_parts(matrix, col_index, header_start_index, code_row_index)
        if not parts and _text(code):
            parts = [_text(code)]
        columns.append({
            "index": col_index,
            "code": _text(code) or None,
            "parts": parts,
            "is_dimension": _is_dimension_column(parts, code),
        })

    dimension_columns = [col for col in columns if col["is_dimension"]]
    metric_columns = [
        col for col in columns
        if not col["is_dimension"] and col["parts"] and _text(col["code"])
    ]

    records: List[Dict[str, Any]] = []
    for row_index in range(code_row_index + 1, len(matrix)):
        row = matrix[row_index]
        if _is_empty_row(row):
            continue

        dimensions = _dimension_values(matrix, row_index, dimension_columns)
        if not dimensions["row_name"] and not any(dimensions["extra_values"]):
            continue

        for metric_col in metric_columns:
            raw_value = _cell(matrix, row_index, metric_col["index"])
            if raw_value is None or _text(raw_value) == "":
                continue

            metric_path = _join_path(metric_col["parts"])
            record = _base_record(meta)
            record.update({
                "row_name": dimensions["row_name"],
                "row_code": dimensions["row_code"],
                "dimension_path": _join_path(dimensions["extra_values"]),
                "dimension_1": _list_get(dimensions["extra_values"], 0),
                "dimension_2": _list_get(dimensions["extra_values"], 1),
                "dimension_3": _list_get(dimensions["extra_values"], 2),
                "dimension_4": _list_get(dimensions["extra_values"], 3),
                "metric_code": metric_col["code"],
                "metric_name": metric_path,
                "metric_path": metric_path,
                "metric_value": _text(raw_value),
                "metric_number": _to_number(raw_value),
                "unit": dimensions["unit"],
            })
            records.append(record)
    return records


def _parse_key_value_table(
    matrix: List[List[Any]],
    meta: Dict[str, Optional[str]],
) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    for row_index, row in enumerate(matrix[2:], start=2):
        values = [_text(value) for value in row]
        non_empty = [value for value in values if value]
        if len(non_empty) < 2:
            continue

        row_code = values[0] or None
        row_name = values[1] if len(values) > 1 else None
        if not row_name or REPORT_CODE_LABEL in row_name or STAT_TIME_LABEL in row_name:
            continue

        metric_values = _dedupe([value for value in values[2:] if value])
        for col_index, value in enumerate(metric_values, start=1):
            if not value:
                continue
            record = _base_record(meta)
            record.update({
                "row_name": row_name,
                "row_code": row_code,
                "dimension_path": None,
                "dimension_1": None,
                "dimension_2": None,
                "dimension_3": None,
                "dimension_4": None,
                "metric_code": str(col_index),
                "metric_name": row_name,
                "metric_path": row_name,
                "metric_value": value,
                "metric_number": _to_number(value),
                "unit": None,
            })
            records.append(record)
    return records


def _cell(matrix: List[List[Any]], row_index: int, col_index: int) -> Any:
    if row_index < 0 or row_index >= len(matrix):
        return None
    row = matrix[row_index]
    if col_index < 0 or col_index >= len(row):
        return None
    return row[col_index]


def _header_parts(
    matrix: List[List[Any]],
    col_index: int,
    header_start_index: int,
    code_row_index: int,
) -> List[str]:
    parts = []
    for row_index in range(header_start_index, code_row_index):
        value = _text(_cell(matrix, row_index, col_index))
        if value:
            parts.append(value)
    return _dedupe(parts)


def _is_dimension_column(parts: List[str], code: Any) -> bool:
    code_text = _text(code)
    header_text = " ".join(parts)
    if code_text in CHINESE_CODE_MARKERS:
        return True
    if any(term in header_text for term in DIMENSION_TERMS):
        return True
    if code_text and not re.fullmatch(r"\d+", code_text):
        return True
    return False


def _dimension_values(
    matrix: List[List[Any]],
    row_index: int,
    dimension_columns: List[Dict[str, Any]],
) -> Dict[str, Any]:
    row_name = None
    row_code = None
    unit = None
    extra_values: List[str] = []

    for column in dimension_columns:
        value = _text(_cell(matrix, row_index, column["index"]))
        if not value:
            continue

        header_text = " ".join(column["parts"])
        if "\u4ee3\u7801" in header_text and row_code is None:
            row_code = value
            continue
        if "\u8ba1\u91cf\u5355\u4f4d" in header_text or header_text == "\u5355\u4f4d":
            unit = value
            continue
        if row_name is None and (
            "\u6307\u6807" in header_text
            or "\u4e13\u4e1a" in header_text
            or "\u540d\u79f0" in header_text
        ):
            row_name = value
            continue
        extra_values.append(value)

    if row_name is None and extra_values:
        row_name = extra_values.pop(0)

    extra_values = _dedupe(extra_values)
    return {
        "row_name": row_name,
        "row_code": row_code,
        "unit": unit,
        "extra_values": extra_values,
    }


def _base_record(meta: Dict[str, Optional[str]]) -> Dict[str, Any]:
    return {
        "source_sheet": meta.get("source_sheet"),
        "report_code": meta.get("report_code"),
        "report_name": meta.get("report_name"),
        "stat_period": meta.get("stat_period"),
        "stat_year": meta.get("stat_year"),
    }


def _is_empty_row(row: List[Any]) -> bool:
    return not any(_text(value) for value in row)


def _join_path(values: List[Optional[str]]) -> Optional[str]:
    cleaned = [value for value in _dedupe([_text(value) for value in values]) if value]
    return "/".join(cleaned) if cleaned else None


def _dedupe(values: List[str]) -> List[str]:
    result = []
    for value in values:
        text = _text(value)
        if text and (not result or result[-1] != text):
            result.append(text)
    return result


def _list_get(values: List[str], index: int) -> Optional[str]:
    return values[index] if index < len(values) else None


def _to_number(value: Any) -> Optional[float]:
    text = _text(value)
    if not text or text in {"-", "--", "\u2014"}:
        return None
    text = text.replace(",", "").replace("\uff0c", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None
